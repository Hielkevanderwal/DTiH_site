# -*- coding: utf-8 -*-
"""
Created on Wed Oct  5 10:56:18 2022

@author: 20212077
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import heartpy as hp
import openpyxl

SCORE_MEDIAN = {'rmssd': (45,27), 'pnn50': (0.11,0.04), 'lf/hf': (0.88,2.75)}

def read_excel(file_name,column_nr=1,array=False):
    """This function reads the heart rate signal data from an excel file
    and saves it into a list. If array = True, then the data will be saved
    into an array, which can be used in HeartPy.
    """
    wb = openpyxl.load_workbook(filename = file_name)
    ws = wb.active
    nr_rows = ws.max_row
    
    timestamp = []
    ppg_signal = []
    for row in range(1,nr_rows+1):
        timestamp.append(ws.cell(row=row,column=column_nr).value)
        ppg_signal.append(ws.cell(row=row,column=column_nr+1).value)
    
    if array:
        timestamp = np.array(timestamp)
        ppg_signal = np.array(ppg_signal)
    
    return timestamp,ppg_signal

def plot_PPG_signal(timestamp,ppg_signal):
    """This function plots the heart rate signal data using matplotlib. The data
    must be stored in a list.
    
    Data: PPG-signal measurements stored in a list.
    """
    plt.figure(figsize=(12,6))
    plt.plot(timestamp,ppg_signal)
    plt.title('Heart rate signal (PPG)')
    plt.xlabel('Time (s)')
    plt.legend(['Heart rate signal'])
    plt.show()

def preprocessing_ppg_signal(data,sample_rate):
    """This function preprocesses the heart rate signal so that it can be
    analyzed by HeartPy.
    
    Data: PPG-signal measurements stored in a numpy array. 
    Sample_rate: the amount of measurements per second (integer).
    """
    data = np.nan_to_num(data)
    filtered_ppg = hp.filter_signal(data, cutoff = [0.8,2.5],
                                    filtertype = 'bandpass',
                                    sample_rate = sample_rate, 
                                    order = 3,
                                    return_top = False)
    
    return filtered_ppg     

def peak_detection(filtered_data,sample_rate):
    """This function analyzes the heart rate signal and detects the peaks. It also
    calculates interesting parameters about the signal and returns them in a dictionary. 
    
    Data: PPG-signal measurements stored in a numpy array. 
    Sample_rate: the amount of measurements per second (integer).
    """
    wd,m = hp.process(filtered_data, sample_rate=sample_rate, calc_freq=True)
    
    return wd,m
     
def plot_peak_detection(peak_data,sample_rate):
    """This function plots the heart rate signal with peak detection and prints the
    calculated parameters.
    """
    hp.plotter(peak_data[0],peak_data[1])

    for measure in peak_data[1]:
        print('{}: {}'.format(measure,peak_data[1][measure]))
    
def poincare_plot(peak_data,sample_rate):
    """This function plots the Poincare plot of the heart rate signal. 
    
    Data: PPG-signal measurements stored in a numpy array. 
    Sample_rate: the amount of measurements per second (integer).
    """
    hp.plot_poincare(peak_data[0],peak_data[1])

def counting_score1(peak_data,score_interval):
    """This function checks in which interval the relevant variables fall: AF or non AF. These 
    intervals are based on literature."""
    score_AF = 0
    score_non_AF = 0
    
    for measure,value in score_interval.items():
        if peak_data[1][measure] >= value[0][0] and peak_data[1][measure] <= value[0][1] and peak_data[1][measure] >= value[1][0] and peak_data[1][measure] <= value[1][1]:
            score_AF =+ 0.5
            score_non_AF =+ 0.5
        elif peak_data[1][measure] >= value[0][0] and peak_data[1][measure] <= value[0][1]:
            score_AF =+ 1
        elif peak_data[1][measure] >= value[1][0] and peak_data[1][measure] <= value[1][1]:
            score_non_AF =+ 1
    
    return score_AF,score_non_AF
    
def counting_score2(peak_data,score_median):
    """This functions determines the category of the parameter (AF or Non-AF) based on
    the median found in the literature."""
    score_AF = 0
    score_non_AF = 0
    
    for measure,median in score_median.items():
        if abs(peak_data[1][measure] - median[0]) < abs(peak_data[1][measure] - median[1]): # AF median is closer
            score_AF += 1
        else:
            score_non_AF += 1
    
    return score_AF,score_non_AF

# def counting_score3(df_input: pd.DataFrame) -> pd.DataFrame:

#     for predictor in ['rmssd', 'pnn50', 'lf/hf']:
#         df_input["normalized_" + predictor] = df_input[predictor].apply(lambda x: interpolation(x, predictor), )
    

#     return df_input


def collect_all_data(total_patients,movement,nr_measurement):
    """This function gathers all the data from multiple patients"""
    rmssd = []
    pnn50 = []
    lfhf = []
    for i in range(1,total_patients+1):
        timestamp,ppg_signal = read_excel('s'+str(i)+'_'+movement+str(nr_measurement)+'_ppg.xlsx',1,True)
        filtered_ppg_signal = preprocessing_ppg_signal(ppg_signal,400)
        peak_data = peak_detection(filtered_ppg_signal,400) # wd,m!
        rmssd.append(peak_data[1]['rmssd'])
        pnn50.append(peak_data[1]['pnn50'])
        lfhf.append(peak_data[1]['lf/hf'])   
      
    return rmssd,pnn50,lfhf

if __name__ == '__main__':
    #timestamp,ppg_signal = read_excel('s1_step1_ppg.xlsx',1,True)
    #filtered_ppg_signal = preprocessing_ppg_signal(ppg_signal,400)
    #peak_data = peak_detection(filtered_ppg_signal,400) # wd,m!

    # Score interval: {'parameter': [(AF occurrence interval),(Non AF occurrence interval)]}
    score_interval = {'rmssd': [(24,108),(19,41)], 'pnn50': [(0.02,0.36),(0.01,0.11)], 'hf': [(7.6,23),(4.8,16.6)]}
    # Score median: {'parameter': (AF median, Non AF median)}

    #plot_PPG_signal(timestamp,ppg_signal)
    #plot_peak_detection(peak_data,400)
    #poincare_plot(peak_data,400)
    #print(counting_score2(peak_data,score_median))

    print(collect_all_data(2,'rest',1))
